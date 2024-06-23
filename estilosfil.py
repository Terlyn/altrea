# -*- coding: utf-8 -*-
import numpy as np
import pandas as pd
import openpyxl
from socal_jaccard import socal_jaccard
from coorp3c import coorp3c
from puntos_ideales import puntos_ideales


# Cargar los datos desde el archivo de Excel
basedatos = '/Users/macbookpro15/Documents/II_PAC_2023/Programación_Comercial/Proyecto/BackEnd/Base/EstilosInteligencias.xlsx'
celdasestilos = 'C1:CD5'
hoja_index = 2  # Índice numérico de la hoja (comenzando desde 0)

wb = openpyxl.load_workbook(basedatos, read_only=True)
hojas = wb.sheetnames

sheet_name = hojas[hoja_index]
sheet = wb[sheet_name]

rows = sheet[celdasestilos]

estilosideales = []
for row in rows:
    estilosideales.append([cell.value for cell in row])

estilosideales = np.array(estilosideales[1:])  # Eliminar la primera fila de encabezados si es necesario


# Calcular las distancias
S_Sokal, S_Jaccard, D2_S, D2_J = socal_jaccard(estilosideales)

# Calcular las coordenadas principales
R, vaps, percent, acum = coorp3c(D2_S)


celdasestilos45 = 'B2:CC10'
hojas = wb.sheetnames

sheet_name = hojas[3]
sheet = wb[sheet_name]

rows = sheet[celdasestilos45]

estu45 = []

for row in rows:
    estu45.append([cell.value for cell in row])


estu45 = np.array(estu45[1:])


# Calcular los puntos ideales
X = estilosideales
x = estu45
m, n = x.shape
r = np.zeros((m, 3))
yt = np.zeros((8, 3))

print(X)
print(x)

for i in range(m):
    yt[i] = puntos_ideales(X, x[i, :])
    r[i, 0:3] = yt[i:i+1][0]

# print(yt)
# print(r)

p, q = r.shape
n, m = R.shape

normat = np.zeros((p, n))
for j in range(p):
    for i in range(n):
        normat[j, i] = np.linalg.norm(R[i, :] - r[j, :])

print(normat)
